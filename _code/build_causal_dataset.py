#!/usr/bin/env python3
"""把 Volve 官方生产 xlsx 加工成 analysis-ready 因果产能预测数据集。

输出(写到 _data/volve_causal_v0.1/):
  daily_production.csv   每井每天一行,列名与原 xlsx 一致,附因果角色注释见 README
  monthly_production.csv 月度表
  well_metadata.csv      7 个井眼的角色、生产/注水区间、转注日期(天然实验锚点)
数据许可: Equinor Open Data Licence(见 license.txt),再分发须署名 Equinor。
"""
import csv
import datetime as dt
from pathlib import Path

import openpyxl

BASE = Path(__file__).resolve().parent.parent / "_data" / "volve_causal_v0.1"
XLSX = BASE / "Volve production data.xlsx"


def dump_sheet(ws, out_path):
    with open(out_path, "w", newline="") as f:
        w = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            w.writerow(["" if v is None else v for v in row])


def build_well_metadata(ws):
    """逐井统计角色区间;干预角色以 FLOW_KIND(production/injection)逐日为准。"""
    wells = {}
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: header.index(name) for name in (
        "DATEPRD", "NPD_WELL_BORE_NAME", "BORE_OIL_VOL", "BORE_WI_VOL", "FLOW_KIND", "WELL_TYPE")}
    for r in ws.iter_rows(min_row=2, values_only=True):
        name = r[idx["NPD_WELL_BORE_NAME"]]
        d = r[idx["DATEPRD"]]
        if isinstance(d, dt.datetime):
            d = d.date()
        s = wells.setdefault(name, {
            "first": d, "last": d, "days": 0, "oil_days": 0,
            "wi_days": 0, "types": set(), "first_wi": None, "last_oil": None})
        s["days"] += 1
        s["first"], s["last"] = min(s["first"], d), max(s["last"], d)
        s["types"].add(str(r[idx["WELL_TYPE"]]))
        if r[idx["BORE_OIL_VOL"]] not in (None, 0, "0"):
            s["oil_days"] += 1
            s["last_oil"] = max(s["last_oil"], d) if s["last_oil"] else d
        if r[idx["BORE_WI_VOL"]] not in (None, 0, "0"):
            s["wi_days"] += 1
            s["first_wi"] = min(s["first_wi"], d) if s["first_wi"] else d
    rows = []
    for name, s in sorted(wells.items()):
        role = "producer" if s["wi_days"] == 0 and s["oil_days"] > 0 else (
            "injector" if s["oil_days"] == 0 else "producer_to_injector")
        rows.append({
            "well": name, "role": role,
            "record_start": s["first"], "record_end": s["last"],
            "record_days": s["days"], "oil_producing_days": s["oil_days"],
            "injection_days": s["wi_days"],
            "first_injection_date": s["first_wi"] or "",
            "last_oil_date": s["last_oil"] or "",
            "well_types_seen": "|".join(sorted(s["types"])),
        })
    return rows


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    daily = wb["Daily Production Data"]
    dump_sheet(daily, BASE / "daily_production.csv")
    dump_sheet(wb["Monthly Production Data"], BASE / "monthly_production.csv")
    meta = build_well_metadata(wb["Daily Production Data"])
    with open(BASE / "well_metadata.csv", "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(meta[0].keys()))
        w.writeheader()
        w.writerows(meta)
    print(f"daily rows={daily.max_row - 1}, wells={len(meta)}")
    for m in meta:
        print(" ", m["well"], m["role"], m["first_injection_date"] or "-")


if __name__ == "__main__":
    main()
